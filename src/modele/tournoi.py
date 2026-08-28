import json
import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
from openpyxl.worksheet.table import Table

from src.modele.equipe import Equipe
from src.modele.joueur import Joueur
from src.modele.partie import Partie, ResultatMatch
from src.vue.style import Couleur


class Tournoi:
    nom: str  # Le nom du tournoi général
    parties: list[Partie]  # La liste des parties qui ont été disputées durant le tournoi
    equipes: dict[str, Equipe]  # La liste des équipes inscrites

    def __init__(self, nom: str = "Tournoi"):
        self.nom = nom
        self.parties = []
        self.equipes = {}

    def ajouter_partie(self, partie: Partie):
        self.parties.append(partie)

    def ajouter_partie_depuis_resultat(self, resultat: ResultatMatch):
        self.ajouter_partie(Partie.from_resultat(resultat))
        self.equipes[resultat.scores["Équipe"]["nom_A"]].ajouter_partie(resultat)
        self.equipes[resultat.scores["Équipe"]["nom_B"]].ajouter_partie(resultat)

    def ajouter_equipe(self, equipe: Equipe):
        if equipe.code not in self.equipes:
            self.equipes[equipe.code] = equipe

    def trier_parties(self):
        parties_niveau: list[list[Partie]] = [[], [], [], [], []]
        for partie in self.parties:
            parties_niveau[partie.niveau - 1].append(partie)
        return parties_niveau

    def trier_equipes(self):
        equipes_niveau: list[list[Equipe]] = [[], [], [], [], []]
        for code in self.equipes:
            equipes_niveau[int(code[0]) - 1].append(self.equipes[code])
        return equipes_niveau

    def classement_niveau(self, niveau: int) -> list[Equipe]:
        equipes = self.trier_equipes()[niveau - 1]

        def cle_tri(equipe: Equipe):
            code = equipe.code

            critere1 = self.points(code) / self.pj(code) if self.pj(code) > 0 else 0
            critere2 = self.moy_pts_diff(code)
            critere3 = equipe.moy_pts_pour()

            return critere1, critere2, critere3

        return sorted(equipes, key=cle_tri, reverse=True)

    # Gestion d'une équipe pour les stats
    def pj(self, code: str) -> int:
        equipe = self.equipes[code]
        return len([p for p in self.parties if equipe.a_joue_partie(p)])

    def victoires(self, code: str) -> int:
        equipe = self.equipes[code]
        return sum(
            equipe.code == partie.vainqueur
            for partie in self.parties
        )

    def defaites(self, code: str) -> int:
        equipe = self.equipes[code]
        return sum(
            (
                    (partie.vainqueur is not None and partie.vainqueur != equipe.code)
                    or
                    (partie.vainqueur is None and equipe.scores[partie.numero - 1] is None)
            )
            for partie in self.parties
            if equipe.a_joue_partie(partie)
        )

    def nulles(self, code: str) -> int:
        equipe = self.equipes[code]
        return sum(
            partie.vainqueur is None
            for partie in self.parties
            if equipe.scores[partie.numero - 1] is not None
            and equipe.a_joue_partie(partie)
        )

    def points(self, code: str) -> int:
        return 2 * self.victoires(code) + self.nulles(code)

    def moy_pts_diff(self, code: str):
        equipe = self.equipes[code]
        total = 0
        pj = 0

        for partie in self.parties:
            if not equipe.a_joue_partie(partie):
                continue
            score_equipe = equipe.scores[partie.numero - 1]
            adv_code = partie.eq_a if equipe.code == partie.eq_b else partie.eq_b
            score_adversaire = self.equipes[adv_code].scores[partie.numero - 1]

            if score_equipe is not None and score_adversaire is not None:
                total += score_equipe - score_adversaire
                pj += 1

        return total / pj if pj > 0 else 0

    # Stockage vers un fichier
    def to_dict(self):
        return {
            "nom": self.nom,
            "parties": [p.to_dict() for p in self.parties],
            "equipes": {code: eq.to_dict() for code, eq in self.equipes.items()}
        }

    @staticmethod
    def from_dict(data: dict) -> 'Tournoi':
        t = Tournoi(data["nom"])
        t.parties = [Partie.from_dict(p) for p in data["parties"]]
        t.equipes = {code: Equipe.from_dict(eq) for code, eq in data["equipes"].items()}
        return t

    def sauvegarder_json(self, chemin: str) -> None:
        with open(chemin, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, ensure_ascii=False, indent=2)

    def generer_excel(self, chemin: str):
        def creer_page_classement(wb: Workbook):
            ws = wb.create_sheet(title="Classement Général")

            def creer_tableau_niveau(niveau: int, r_init: int):
                """
                La fonction sert à créer un tableau pour afficher le classement dans le fichier Excel
                :param niveau: Le niveau dont on fait le classement.
                :param r_init: La rangée de début du tableau.
                :return: La rangée de fin du tableau.
                """
                # -----------------------------
                #         Style
                # -----------------------------
                NOIR = "FF000000"
                BLANC = "FFFFFF"
                couleurs = {
                    1: Couleur.ORANGE[1:],
                    2: Couleur.ROSE[1:],
                    3: Couleur.BLEU[1:],
                    4: Couleur.VERT[1:],
                    5: Couleur.JAUNE[1:],
                }

                POLICE_TITRE = Font(name="Arial", size=16, bold=True, color=BLANC)
                POLICE_ENTETE = Font(name="Arial", size=11, bold=True, color=NOIR)
                POLICE_DONNEES = Font(name="Arial", size=11, color=NOIR)

                REMPLISSAGE_TITRE = PatternFill("solid", fgColor=NOIR, bgColor=NOIR)
                REMPLISSAGE_ENTETE = PatternFill("solid", fgColor=couleurs[niveau])
                REMPLISSAGE_CODE = PatternFill("solid", fgColor=couleurs[niveau])

                BORDURE_FINE = Side(style="thin", color=NOIR)
                BORDURE_EPAISSE = Side(style="medium", color=NOIR)
                PAS_BORDURE = Side(style=None)

                def bordure_cellule(rangee, col_idx, r_titre, r_bas, nb_colonnes):
                    r_entete = r_titre + 1
                    est_coloree = rangee in (r_titre, r_entete) or col_idx in (1, 2)

                    haut = bas = gauche = droite = BORDURE_FINE
                    if est_coloree:
                        haut = bas = gauche = droite = PAS_BORDURE

                    if rangee == r_titre:
                        haut = BORDURE_EPAISSE
                    if rangee == r_bas:
                        bas = BORDURE_EPAISSE
                    if col_idx == 1:
                        gauche = BORDURE_EPAISSE
                    if col_idx == nb_colonnes:
                        droite = BORDURE_EPAISSE

                    return Border(left=gauche, right=droite, top=haut, bottom=bas)

                CENTRE = Alignment(horizontal="center", vertical="center")
                GAUCHE = Alignment(horizontal="left", vertical="center")

                # --- Colonnes du tableau : (en-tête, format nombre, alignement) ---
                COLONNES = [
                    ("Rang", "0", CENTRE),
                    ("Code", "0", CENTRE),
                    ("Nom de l'équipe", "0", GAUCHE),
                    ("# Parties", "0", CENTRE),
                    ("V", "0", CENTRE),
                    ("D", "0", CENTRE),
                    ("N", "0", CENTRE),
                    ("Points", "0", CENTRE),
                    ("Moy pts/match", "0.00", CENTRE),
                    ("Pts +", "0", CENTRE),
                    ("Pts -", "0", CENTRE),
                    ("Diff +/-", "0", CENTRE),
                    ("Moy +/-", "0.00", CENTRE),
                    ("Moy pts +", "0.00", CENTRE),
                ]

                equipes_classees = self.classement_niveau(niveau)

                # -------- Ligne de titre --------
                ws.merge_cells(start_row=r_init, start_column=1, end_row=r_init, end_column=len(COLONNES))
                cell_titre = ws.cell(row=r_init, column=1, value=f'Secondaire {niveau}')
                cell_titre.font = POLICE_TITRE
                cell_titre.fill = REMPLISSAGE_TITRE
                cell_titre.alignment = CENTRE
                ws.row_dimensions[1].height = 24

                # -------- En-têtes --------
                for idx_col, (entete, fmt, align) in enumerate(COLONNES):
                    cell = ws.cell(row=(r_init + 1), column=idx_col + 1, value=entete)
                    cell.font = POLICE_ENTETE
                    cell.fill = REMPLISSAGE_ENTETE
                    cell.alignment = CENTRE
                    cell.border = cell.border = bordure_cellule(r_init, idx_col, r_init + 1,
                                                                r_init + len(equipes_classees) + 1,
                                                                len(COLONNES))

                    # --- Lignes de données ---
                for rang, equipe in enumerate(equipes_classees, start=1):
                    code = equipe.code
                    nb_matchs = self.pj(code)
                    victoires = self.victoires(code)
                    defaites = self.defaites(code)
                    nulles = self.nulles(code)
                    points = self.points(code)
                    moy_pts_match = points / nb_matchs if nb_matchs else 0
                    pts_pour = sum(s for s in equipe.scores if s is not None)
                    pts_contre = sum(
                        self.equipes[
                            p.eq_a if code == p.eq_b else p.eq_b
                        ].scores[p.numero - 1] or 0
                        for p in self.parties
                        if equipe.a_joue_partie(p) and equipe.scores[p.numero - 1] is not None
                    )
                    diff = pts_pour - pts_contre
                    moy_diff = self.moy_pts_diff(code)
                    moy_pts_pour = equipe.moy_pts_pour()

                    valeurs = [
                        rang, code, equipe.nom, nb_matchs, victoires, defaites, nulles,
                        points, moy_pts_match, pts_pour, pts_contre, diff, moy_diff, moy_pts_pour,
                    ]

                    rangee = r_init + rang + 1
                    for col_idx, (valeur, (_entete, fmt, align)) in enumerate(zip(valeurs, COLONNES), start=1):
                        cell = ws.cell(row=rangee, column=col_idx, value=valeur)
                        cell.font = POLICE_DONNEES
                        cell.alignment = align
                        cell.border = cell.border = bordure_cellule(rangee, col_idx, r_init,
                                                                    r_init + len(equipes_classees) + 1,
                                                                    len(COLONNES))
                        if fmt:
                            cell.number_format = fmt
                        if col_idx in (1, 2):
                            cell.fill = REMPLISSAGE_CODE
                return r_init + len(equipes_classees)

            # Modification des largeurs des colonnes
            largeurs = {
                "A": 7.125,
                "B": 5.375,
                "C": 36.75,
                "D": 8,
                "E": 6.875,
                "F": 6.875,
                "G": 6.875,
                "H": 11,
                "I": 13.5,
                "J": 8.875,
                "K": 8.875,
                "L": 8.875,
                "M": 11,
                "N": 16,
            }

            for col, l in largeurs.items():
                ws.column_dimensions[col].width = l

            ws.row_dimensions[1].height = 36

            # Titre
            ws.merge_cells('E1:K1')
            c = ws.cell(row=1, column=5)
            c.value = self.nom
            c.font = Font(name='Calibri', bold=True, size=24)

            # Classements
            r_debut = 3
            for niveau in self.trier_equipes():
                if len(niveau) > 0:
                    r_debut = creer_tableau_niveau(int(niveau[0].code[0]), r_debut)
                    r_debut += 3

        def creer_page_stats_ind(equipes_niveau: list[Equipe], wb: Workbook):
            def _creer_ligne_joueur(code, nom_equipe, joueur: Joueur, rangee: int, nombre_total_joueurs: int):
                ligne = [code, nom_equipe, joueur.nom, f"=RANK(E{rangee};$E$2:$E${nombre_total_joueurs + 1})",
                         joueur.moyenne()]
                for score in joueur.scores:
                    if score is not None:
                        ligne.append(score)
                    else:
                        ligne.append("")
                return ligne

            niveau = int(equipes_niveau[0].code[0])

            suffixe_ordinal = "re" if niveau == 1 else "e"
            ws = wb.create_sheet(title=f"Stats ind. - {niveau}{suffixe_ordinal} Sec.")

            # Obtention des équipes et création du tableau
            entetes = ["Code", "Équipe", "Nom", "Rang", "Moyenne"]
            for i in range(20):
                entetes.append(f"Match{i + 1}")
            ws.append(entetes)  # En-têtes

            # Ajout des données des joueurs au tableau
            total_joueurs = sum(len(eq.joueurs) for eq in equipes_niveau)
            i = 2
            for equipe in equipes_niveau:
                for joueur in equipe.joueurs:
                    ws.append(_creer_ligne_joueur(equipe.code, equipe.nom, joueur, i, total_joueurs))
                    i += 1

            tab = Table(displayName=f"StatsIndSec{niveau}", ref=f"A1:Y{i}")

            ws.add_table(tab)

        def creer_page_resultats_match(equipes_niveau: list[Equipe], wb: Workbook):
            niveau = int(equipes_niveau[0].code[0])
            parties = self.trier_parties()[niveau - 1]
            if not parties:
                return

            # -----------------------------
            #         Style
            # -----------------------------
            NOIR = "FF000000"
            couleurs = {
                1: Couleur.ORANGE[1:],
                2: Couleur.ROSE[1:],
                3: Couleur.BLEU[1:],
                4: Couleur.VERT[1:],
                5: Couleur.JAUNE[1:],
            }

            POLICE_ENTETE = Font(name="Arial", size=11, bold=True)
            POLICE_DONNEES = Font(name="Arial", size=11)
            REMPLISSAGE_CODE = PatternFill("solid", fgColor=couleurs[niveau])
            BORDURE_EPAISSE = Side(style="medium", color=NOIR)
            PAS_BORDURE = Side(style=None)
            CENTRE = Alignment(horizontal="center", vertical="center")

            def bordure_code(est_premiere_rangee, est_derniere_rangee):
                return Border(
                    left=BORDURE_EPAISSE, right=PAS_BORDURE,
                    top=BORDURE_EPAISSE if est_premiere_rangee else PAS_BORDURE,
                    bottom=BORDURE_EPAISSE if est_derniere_rangee else PAS_BORDURE,
                )

            def bordure_score(est_premiere_rangee, est_derniere_rangee):
                return Border(
                    left=PAS_BORDURE, right=BORDURE_EPAISSE,
                    top=BORDURE_EPAISSE if est_premiere_rangee else PAS_BORDURE,
                    bottom=BORDURE_EPAISSE if est_derniere_rangee else PAS_BORDURE,
                )

            suffixe_ordinal = "re" if niveau == 1 else "e"
            ws = wb.create_sheet(title=f"Résultats - Matchs {niveau}{suffixe_ordinal} sec.")

            # Regroupement des parties par plateau
            parties_plateau: dict[int, dict[int, Partie]] = {}
            for partie in parties:
                parties_plateau.setdefault(partie.plateau, {})[partie.numero] = partie

            match_max = max(p.numero for p in parties)
            plateau_min = min(parties_plateau)
            plateau_max = max(parties_plateau)

            # En-têtes des matchs (fusionnées sur les 2 colonnes code/score de chaque match)
            for match in range(1, match_max + 1):
                col_code = 2 + (match - 1) * 2
                cell = ws.cell(row=1, column=col_code, value=f"M{match}")
                cell.font = POLICE_ENTETE
                cell.alignment = CENTRE
                ws.merge_cells(start_row=1, start_column=col_code, end_row=1, end_column=col_code + 1)

            # Corps du tableau : 2 lignes par plateau (équipe A, puis équipe B).
            ligne = 2
            for plateau in range(plateau_min, plateau_max + 1):
                label = ws.cell(row=ligne, column=1, value=f"Plateau {plateau}")
                label.font = POLICE_ENTETE
                label.alignment = CENTRE
                ws.merge_cells(start_row=ligne, start_column=1, end_row=ligne + 1, end_column=1)

                parties_du_plateau = parties_plateau.get(plateau, {})

                for match in range(1, match_max + 1):
                    col_code = 2 + (match - 1) * 2
                    col_score = col_code + 1
                    partie = parties_du_plateau.get(match)

                    if partie is not None:
                        equipe_a = self.equipes[partie.eq_a]
                        equipe_b = self.equipes[partie.eq_b]
                        score_a = equipe_a.scores[match - 1]
                        score_b = equipe_b.scores[match - 1]

                        if score_a is None or score_b is None:
                            if partie.vainqueur == partie.eq_a:
                                score_a, score_b = "Victoire", "Forfait"
                            elif partie.vainqueur == partie.eq_b:
                                score_a, score_b = "Forfait", "Victoire"
                            else:
                                score_a = score_b = "Forfait"
                        paires = [(int(partie.eq_a), score_a), (int(partie.eq_b), score_b)]
                    else:
                        paires = [(None, None), (None, None)]

                    for decalage, (code, score) in enumerate(paires):
                        rangee = ligne + decalage
                        premiere = decalage == 0
                        derniere = decalage == 1

                        c_code = ws.cell(row=rangee, column=col_code, value=code)
                        c_score = ws.cell(row=rangee, column=col_score, value=score)

                        c_code.font = c_score.font = POLICE_DONNEES
                        c_code.alignment = c_score.alignment = CENTRE
                        c_code.border = bordure_code(premiere, derniere)
                        c_score.border = bordure_score(premiere, derniere)
                        c_code.fill = REMPLISSAGE_CODE

                ligne += 2

            # Largeurs des colonnes
            ws.column_dimensions["A"].width = 12
            for col in range(2, 2 + match_max * 2):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

        # Création du fichier
        wb = Workbook()
        wb.remove(wb.active)

        creer_page_classement(wb)

        for niveau in self.trier_equipes():
            if len(niveau) > 0:
                creer_page_resultats_match(niveau, wb)
                creer_page_stats_ind(niveau, wb)

        wb.save(os.path.abspath(chemin))

    @staticmethod
    def charger(chemin: str):
        if chemin.endswith('.xlsx'):

            tournoi = Tournoi()

            equipes_locales: dict[str, Equipe] = {}
            wb = openpyxl.load_workbook(chemin)
            ws = wb.active
            if ws is None:
                return equipes_locales

            nb_rangees = ws.max_row or 0
            compteurs = [0, 0, 0, 0, 0]

            compte_ecoles: dict[tuple[str, str], int] = {}
            for i in range(2, nb_rangees + 1):
                ecole_val = ws.cell(row=i, column=1).value
                ecole = str(ecole_val).strip().title() if ecole_val is not None else ""
                niveau_val = ws.cell(row=i, column=2).value
                niveau = str(niveau_val)[0] if niveau_val is not None else "0"
                cle = (niveau, ecole)
                compte_ecoles[cle] = compte_ecoles.get(cle, 0) + 1

            compteur_lettre: dict[tuple[str, str], int] = {}

            for i in range(2, nb_rangees + 1):
                # Obtention de l'école
                ecole_val = ws.cell(row=i, column=1).value
                ecole = str(ecole_val).strip().title() if ecole_val else ""

                # Obtention du niveau et création du code
                niveau_val = ws.cell(row=i, column=2).value
                niveau = str(niveau_val)[0] if niveau_val else "0"
                compteurs[int(niveau) - 1] += 1

                code = f'{int(niveau) * 100 + compteurs[int(niveau) - 1]}'

                # Création du nom
                cle = (niveau, ecole)
                total_pour_ecole = compte_ecoles[cle]

                if total_pour_ecole > 1:
                    i_lettre = compteur_lettre.get(cle, 0)
                    lettre = chr(ord('A') + i_lettre)
                    compteur_lettre[cle] = i_lettre + 1
                else:
                    lettre = ""

                nom = f'{ecole} {lettre}'.strip()

                j = 4
                joueurs: list[Joueur] = []
                while True:
                    cell_val = ws.cell(row=i, column=j).value
                    if not cell_val:
                        break
                    joueurs.append(Joueur(str(cell_val)))
                    j += 1

                equipes_locales[code] = Equipe(code, nom, joueurs)

            tournoi.equipes = equipes_locales
            return tournoi

        else:
            with open(chemin, 'r', encoding='utf-8') as f:
                return Tournoi.from_dict(json.load(f))
